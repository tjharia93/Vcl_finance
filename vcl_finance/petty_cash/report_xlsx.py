"""Petty Cash weekly workbook.

The layout Tanuj asked for: **categories across the top, days down the rows**,
one row per day Sun -> Sat, with the running balance and the opening -> close
reconciliation underneath.

This module is deliberately free of any ``frappe`` import. It takes plain dicts
(a Petty Cash Sheet ``as_dict()``, or the same shape off the REST API) and hands
back workbook bytes, so the Compass download endpoint and the local ``pcreport``
CLI build the identical file from one implementation.
"""
from __future__ import annotations

import io
from datetime import date, datetime, timedelta

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

CATEGORIES = [
    ("TG", "Transport — Goods"),
    ("TE", "Transport — Employee"),
    ("SE", "Spares / Engineering"),
    ("OA", "Office / Admin"),
    ("FD", "Food"),
    ("GP", "Geeprint"),
    ("OT", "Other"),
]
CAT_CODES = [c for c, _ in CATEGORIES]
MISC_KINDS = ["Bike Fuel", "Forklift"]

# Parking one column per vehicle, and wages one per kind, because that is how the
# paper sheet is written and how it gets checked. A single "Parking" column tells
# you the week cost 4,300 and nothing about which car; a single "Wages & Comm."
# hides a commission inside a wage bill.
VEHICLES = ["KAP 466", "KAY 635", "KCB 430", "KBQ 788", "KBT 972"]
PARK_COLS = [f"P {v}" for v in VEHICLES]
# A plate nobody listed still has to land somewhere. Dropping it would make the
# row totals disagree with the sheet, which is the one thing a report may not do.
PARK_OTHER = "P other"
# entry_type on the row -> the column it belongs in. "Wage" reads as Casual on the
# paper sheet and in the accounts (5120.1.4 Casual Wages), so it is labelled that.
WAGE_COLS = {"Wage": "Casual", "Overtime": "Overtime",
             "Piecework": "Piecework", "Commission": "Commission"}
WAGE_ORDER = ["Casual", "Overtime", "Piecework", "Commission"]

OUT_COLS = (CAT_CODES + PARK_COLS + [PARK_OTHER] + MISC_KINDS
            + WAGE_ORDER + ["Loans"])

# ---- VCL brand -------------------------------------------------------------
NAVY, BLUE = "1D2766", "2B3990"
BLUE_LIGHT, BLUE_PALE = "D6DBF5", "EEF0FB"
INK, INK2, BORDER, BG = "1C1E2E", "4A4E6A", "DDE0F0", "F5F6FA"
GREEN, AMBER = "0B6B3A", "B35A00"

MONEY = "#,##0;(#,##0)"
THIN = Side(style="thin", color=BORDER)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


# ------------------------------------------------------------------- helpers
def as_date(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return datetime.strptime(str(v)[:10], "%Y-%m-%d").date() if v else None


def week_days(week_ending):
    """The seven days Sun -> Sat of the week ``week_ending`` closes."""
    we = as_date(week_ending)
    return [we - timedelta(days=i) for i in range(6, -1, -1)]


def _style(cell, *, bold=False, size=10, colour=INK, fill=None, num=None,
           align=None, wrap=False, box=True, italic=False, name="Calibri"):
    cell.font = Font(name=name, bold=bold, size=size, color=colour, italic=italic)
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)
    if num:
        cell.number_format = num
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if box:
        cell.border = BOX
    return cell


def _put(ws, row, col, value, **kw):
    return _style(ws.cell(row=row, column=col, value=value), **kw)


def _page(ws):
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True


def _title(ws, ncols, subtitle, meta=""):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(ncols, 4))
    _put(ws, 1, 1, "VIMIT CONVERTERS LIMITED", bold=True, size=15, colour=NAVY,
         align="left", box=False)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max(ncols, 4))
    _put(ws, 2, 1, subtitle, bold=True, size=11, colour=BLUE, align="left", box=False)
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=max(ncols, 4))
    _put(ws, 3, 1, meta, size=9, colour=INK2, align="left", box=False, name="Courier New")
    ws.row_dimensions[1].height = 21


def _sheet_meta(sheet):
    return (f"Sheet {sheet.get('name')}  ·  Float: {sheet.get('float') or '—'}  ·  "
            f"Custodian: {sheet.get('custodian_name') or '—'}  ·  "
            f"Week {sheet.get('week_no') or '—'}  ·  Status: {sheet.get('status')}  ·  "
            f"Generated {datetime.now():%d %b %Y %H:%M}")


def daily_totals(sheet, days):
    """{day: {column: amount}} for one sheet, cancelled rows excluded."""
    rows = {day: {k: 0.0 for k in OUT_COLS + ["Cash IN"]} for day in days}
    for v in sheet.get("vouchers") or []:
        day = as_date(v.get("txn_date"))
        if v.get("cancelled") or day not in rows:
            continue
        amt = v.get("amount") or 0
        if v.get("cash_in"):
            rows[day]["Cash IN"] += amt
        else:
            rows[day][v.get("category") if v.get("category") in CAT_CODES else "OT"] += amt
    for p in sheet.get("parking_entries") or []:
        day = as_date(p.get("txn_date"))
        if not p.get("cancelled") and day in rows:
            veh = (p.get("vehicle") or "").strip()
            key = f"P {veh}" if f"P {veh}" in PARK_COLS else PARK_OTHER
            rows[day][key] += p.get("amount") or 0
    for m in sheet.get("misc_entries") or []:
        day = as_date(m.get("txn_date"))
        if not m.get("cancelled") and day in rows and m.get("kind") in MISC_KINDS:
            rows[day][m["kind"]] += m.get("amount") or 0
    for w in sheet.get("wages_entries") or []:
        day = as_date(w.get("txn_date"))
        if not w.get("cancelled") and day in rows:
            # An unrecognised entry_type falls to Casual rather than vanishing —
            # the same reason PARK_OTHER exists.
            rows[day][WAGE_COLS.get(w.get("entry_type"), "Casual")] += w.get("amount") or 0
    for l in sheet.get("loan_entries") or []:
        day = as_date(l.get("txn_date"))
        if not l.get("cancelled") and day in rows:
            rows[day]["Loans"] += l.get("amount_issued") or 0
    return rows


# ------------------------------------------------------------- summary sheet
def build_summary(wb, sheet, tab):
    days = week_days(sheet["week_ending"])
    cols = ["Date", "Day"] + OUT_COLS + ["TOTAL OUT", "Cash IN", "Net movement", "Running balance"]
    n = len(cols)
    ws = wb.create_sheet(tab)
    _title(ws, n,
           f"Petty Cash — weekly summary by category · week ending {days[-1]:%d %B %Y} "
           f"({days[0]:%a %d %b} – {days[-1]:%a %d %b %Y})",
           _sheet_meta(sheet))
    rows = daily_totals(sheet, days)

    hdr = 5
    ws.merge_cells(start_row=hdr, start_column=3, end_row=hdr, end_column=2 + len(CAT_CODES))
    _put(ws, hdr, 3, "VOUCHER CATEGORIES", bold=True, size=9, colour=BLUE, fill=BLUE_PALE, align="center")
    ws.merge_cells(start_row=hdr, start_column=3 + len(CAT_CODES), end_row=hdr, end_column=2 + len(OUT_COLS))
    _put(ws, hdr, 3 + len(CAT_CODES), "SUPPORTING REGISTERS", bold=True, size=9,
         colour=BLUE, fill=BLUE_PALE, align="center")
    ws.merge_cells(start_row=hdr, start_column=3 + len(OUT_COLS), end_row=hdr, end_column=n)
    _put(ws, hdr, 3 + len(OUT_COLS), "TOTALS", bold=True, size=9, colour=BLUE, fill=BLUE_PALE, align="center")
    for c in (1, 2):
        _put(ws, hdr, c, "", fill=BLUE_PALE)

    hdr2 = hdr + 1
    for i, label in enumerate(cols, start=1):
        _put(ws, hdr2, i, label, bold=True, size=9, colour=NAVY, fill=BLUE_LIGHT,
             align="center" if i > 2 else "left", wrap=True)
    ws.row_dimensions[hdr2].height = 30

    r = hdr2 + 1
    running = sheet.get("opening_balance") or 0
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=n - 1)
    _put(ws, r, 1, "Opening balance brought forward", bold=True, colour=INK2, align="left", fill=BG)
    for c in range(2, n):
        _style(ws.cell(row=r, column=c), fill=BG)
    _put(ws, r, n, running, bold=True, num=MONEY, align="right", fill=BG,
         colour=AMBER if running < 0 else GREEN, name="Courier New")

    first = r + 1
    for day in days:
        r += 1
        rec = rows[day]
        _put(ws, r, 1, day, num="dd/mm/yyyy", align="center", name="Courier New")
        _put(ws, r, 2, f"{day:%a}", align="center", colour=INK2)
        for i, key in enumerate(OUT_COLS, start=3):
            _put(ws, r, i, rec[key] or None, num=MONEY, align="right", name="Courier New")
        tot_out = sum(rec[k] for k in OUT_COLS)
        cash_in = rec["Cash IN"]
        running += cash_in - tot_out
        _put(ws, r, 3 + len(OUT_COLS), tot_out or None, bold=True, num=MONEY, align="right", name="Courier New")
        _put(ws, r, 4 + len(OUT_COLS), cash_in or None, num=MONEY, align="right",
             colour=GREEN if cash_in else INK, name="Courier New")
        _put(ws, r, 5 + len(OUT_COLS), (cash_in - tot_out) or None, num=MONEY, align="right", name="Courier New")
        _put(ws, r, n, running, bold=True, num=MONEY, align="right", name="Courier New",
             colour=AMBER if running < 0 else INK)
    last = r

    r += 1
    _put(ws, r, 1, "TOTAL", bold=True, colour="FFFFFF", fill=BLUE, align="left")
    _put(ws, r, 2, "", fill=BLUE)
    for i in range(3, n):
        L = get_column_letter(i)
        _put(ws, r, i, f"=SUM({L}{first}:{L}{last})", bold=True, num=MONEY, align="right",
             colour="FFFFFF", fill=BLUE, name="Courier New")
    _put(ws, r, n, f"={get_column_letter(n)}{last}", bold=True, num=MONEY, align="right",
         colour="FFFFFF", fill=BLUE, name="Courier New")
    totals_row = r

    r += 1
    _put(ws, r, 1, "% of cash out", italic=True, size=9, colour=INK2, align="left")
    _put(ws, r, 2, "")
    tot_ref = f"{get_column_letter(3 + len(OUT_COLS))}{totals_row}"
    for i in range(3, 3 + len(OUT_COLS)):
        L = get_column_letter(i)
        _put(ws, r, i, f'=IF({tot_ref}=0,"",{L}{totals_row}/{tot_ref})',
             num="0.0%", align="right", size=9, colour=INK2, name="Courier New")
    for i in range(3 + len(OUT_COLS), n + 1):
        _put(ws, r, i, "")

    r += 2
    _put(ws, r, 1, "RECONCILIATION", bold=True, colour=NAVY, align="left", box=False)
    counted = sheet.get("cash_count_end") or 0
    for label, val in [
        ("Opening balance", sheet.get("opening_balance") or 0),
        ("Less: total cash out", -(sheet.get("total_out") or 0)),
        ("Add: cash in / refunds", sheet.get("total_in") or 0),
        ("Expected closing balance", sheet.get("expected_close") or 0),
        ("Physical cash counted", counted or None),
        ("Variance (counted − expected)", (sheet.get("variance") or 0) if counted else None),
        ("Authorised float", sheet.get("authorised_float") or 0),
    ]:
        r += 1
        bold = label.startswith("Expected")
        fill = BLUE_PALE if bold else None
        _put(ws, r, 1, label, bold=bold, align="left", fill=fill)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=n - 1)
        for c in range(2, n):
            _style(ws.cell(row=r, column=c), fill=fill)
        if val is None:
            _put(ws, r, n, "not entered", italic=True, size=9, colour=INK2, align="right", fill=fill)
        else:
            _put(ws, r, n, val, bold=bold, num=MONEY, align="right", name="Courier New",
                 fill=fill, colour=AMBER if val < 0 else INK)

    ws.column_dimensions["A"].width = 13
    ws.column_dimensions["B"].width = 7
    for i in range(3, n):
        ws.column_dimensions[get_column_letter(i)].width = 13
    ws.column_dimensions[get_column_letter(n)].width = 17
    ws.freeze_panes = f"C{hdr2 + 1}"
    _page(ws)
    return ws


# -------------------------------------------------------------- voucher sheet
def build_vouchers(wb, sheet, tab):
    days = week_days(sheet["week_ending"])
    cols = ["Date", "Day", "#", "Recipient / description", "Cat", "Category",
            "Cash out", "Cash in", "PC", "ETR", "Notes", "Status"]
    ws = wb.create_sheet(tab)
    _title(ws, len(cols),
           f"Petty Cash — voucher register · week ending {days[-1]:%d %B %Y}",
           _sheet_meta(sheet))
    hdr = 5
    for i, label in enumerate(cols, start=1):
        _put(ws, hdr, i, label, bold=True, size=9, colour=NAVY, fill=BLUE_LIGHT,
             align="center" if i in (3, 5, 7, 8, 9, 10, 12) else "left")
    labels = dict(CATEGORIES)
    r = hdr
    for v in sorted(sheet.get("vouchers") or [],
                    key=lambda x: (str(x.get("txn_date") or "9999"), x.get("row_idx") or 0)):
        if not v.get("txn_date") and not (v.get("amount") or 0):
            continue
        r += 1
        day = as_date(v.get("txn_date"))
        cancelled = bool(v.get("cancelled"))
        col = INK2 if cancelled else INK
        _put(ws, r, 1, day, num="dd/mm/yyyy", align="center", colour=col, name="Courier New")
        _put(ws, r, 2, f"{day:%a}" if day else "", align="center", colour=INK2)
        _put(ws, r, 3, v.get("row_idx"), align="center", colour=INK2, size=9)
        _put(ws, r, 4, v.get("recipient") or "", align="left", colour=col, italic=cancelled)
        _put(ws, r, 5, v.get("category") or ("IN" if v.get("cash_in") else ""), align="center", colour=col)
        _put(ws, r, 6, "Cash in / refund" if v.get("cash_in") else labels.get(v.get("category"), ""),
             align="left", colour=INK2, size=9)
        _put(ws, r, 7, None if v.get("cash_in") else (v.get("amount") or 0), num=MONEY,
             align="right", colour=col, name="Courier New")
        _put(ws, r, 8, (v.get("amount") or 0) if v.get("cash_in") else None, num=MONEY,
             align="right", colour=GREEN, name="Courier New")
        _put(ws, r, 9, "Y" if v.get("pc_received") else "", align="center", size=9, colour=INK2)
        _put(ws, r, 10, "Y" if v.get("etr_received") else "", align="center", size=9, colour=INK2)
        _put(ws, r, 11, v.get("notes") or "", align="left", size=9, colour=INK2, wrap=True)
        _put(ws, r, 12, ("CANCELLED — " + (v.get("cancel_remark") or "")) if cancelled else "",
             align="left", size=9, colour=AMBER, italic=True)
    first, last = hdr + 1, r
    r += 1
    _put(ws, r, 1, "TOTAL", bold=True, colour="FFFFFF", fill=BLUE, align="left")
    for i in range(2, len(cols) + 1):
        _put(ws, r, i, "", fill=BLUE)
    for i in (7, 8):
        L = get_column_letter(i)
        # cancelled rows carry text in the Status column and must not be totalled
        _put(ws, r, i, f'=SUMIF($L${first}:$L${last},"",{L}{first}:{L}{last})', bold=True,
             num=MONEY, align="right", colour="FFFFFF", fill=BLUE, name="Courier New")
    _put(ws, r, 4, "(cancelled rows are listed for audit but excluded from these totals)",
         size=8, italic=True, colour="FFFFFF", fill=BLUE, align="left")

    for L, w in zip("ABCDEFGHIJKL", [12, 6, 5, 40, 6, 20, 13, 13, 5, 5, 42, 34]):
        ws.column_dimensions[L].width = w
    ws.freeze_panes = f"A{hdr + 1}"
    if last >= first:
        ws.auto_filter.ref = f"A{hdr}:{get_column_letter(len(cols))}{last}"
    _page(ws)
    return ws


# ------------------------------------------------------------ registers sheet
def build_registers(wb, sheet, tab):
    days = week_days(sheet["week_ending"])
    ws = wb.create_sheet(tab)
    _title(ws, 2 + len(days) + 1,
           f"Petty Cash — supporting registers · week ending {days[-1]:%d %B %Y}",
           _sheet_meta(sheet))

    r = 5
    _put(ws, r, 1, "PARKING", bold=True, size=11, colour=BLUE, align="left", box=False)
    r += 1
    _put(ws, r, 1, "Vehicle", bold=True, size=9, colour=NAVY, fill=BLUE_LIGHT, align="left")
    _put(ws, r, 2, "", fill=BLUE_LIGHT)
    for i, day in enumerate(days, start=3):
        _put(ws, r, i, f"{day:%a}\n{day:%d/%m}", bold=True, size=9, colour=NAVY,
             fill=BLUE_LIGHT, align="center", wrap=True)
    _put(ws, r, 3 + len(days), "Total", bold=True, size=9, colour=NAVY, fill=BLUE_LIGHT, align="center")
    ws.row_dimensions[r].height = 26
    hdr = r

    grid, vehicles = {}, []
    for p in sheet.get("parking_entries") or []:
        if p.get("cancelled") or not p.get("txn_date"):
            continue
        veh = (p.get("vehicle") or "—").strip()
        if veh not in vehicles:
            vehicles.append(veh)
        key = (veh, as_date(p["txn_date"]))
        grid[key] = grid.get(key, 0) + (p.get("amount") or 0)
    for veh in sorted(vehicles):
        r += 1
        _put(ws, r, 1, veh, align="left", name="Courier New")
        _put(ws, r, 2, "")
        for i, day in enumerate(days, start=3):
            _put(ws, r, i, grid.get((veh, day)) or None, num=MONEY, align="right", name="Courier New")
        _put(ws, r, 3 + len(days),
             f"=SUM({get_column_letter(3)}{r}:{get_column_letter(2 + len(days))}{r})",
             bold=True, num=MONEY, align="right", name="Courier New")
    if vehicles:
        r += 1
        _put(ws, r, 1, "Total", bold=True, colour="FFFFFF", fill=BLUE, align="left")
        _put(ws, r, 2, "", fill=BLUE)
        for i in range(3, 4 + len(days)):
            L = get_column_letter(i)
            _put(ws, r, i, f"=SUM({L}{hdr + 1}:{L}{r - 1})", bold=True, num=MONEY,
                 align="right", colour="FFFFFF", fill=BLUE, name="Courier New")

    for kind in MISC_KINDS:
        r += 2
        _put(ws, r, 1, kind.upper(), bold=True, size=11, colour=BLUE, align="left", box=False)
        r += 1
        for i, label in enumerate(["Date", "Day", "Amount", "Notes"], start=1):
            _put(ws, r, i, label, bold=True, size=9, colour=NAVY, fill=BLUE_LIGHT,
                 align="center" if i in (2, 3) else "left")
        first = r + 1
        for m in sheet.get("misc_entries") or []:
            if m.get("kind") != kind or m.get("cancelled") or not m.get("txn_date"):
                continue
            r += 1
            day = as_date(m["txn_date"])
            _put(ws, r, 1, day, num="dd/mm/yyyy", align="center", name="Courier New")
            _put(ws, r, 2, f"{day:%a}", align="center", colour=INK2)
            _put(ws, r, 3, m.get("amount") or 0, num=MONEY, align="right", name="Courier New")
            _put(ws, r, 4, m.get("notes") or "", align="left", size=9, colour=INK2)
        r += 1
        _put(ws, r, 1, "Total", bold=True, colour="FFFFFF", fill=BLUE, align="left")
        _put(ws, r, 2, "", fill=BLUE)
        _put(ws, r, 3, f"=SUM(C{first}:C{r - 1})" if r > first else 0, bold=True, num=MONEY,
             align="right", colour="FFFFFF", fill=BLUE, name="Courier New")
        _put(ws, r, 4, "", fill=BLUE)

    r += 2
    _put(ws, r, 1, "WAGES & COMMISSION", bold=True, size=11, colour=BLUE, align="left", box=False)
    r += 1
    for i, label in enumerate(["Date", "Day", "Amount", "Recipient", "Type", "Reason"], start=1):
        _put(ws, r, i, label, bold=True, size=9, colour=NAVY, fill=BLUE_LIGHT,
             align="center" if i in (2, 3) else "left")
    first = r + 1
    for w in sorted(sheet.get("wages_entries") or [],
                    key=lambda x: (str(x.get("txn_date") or "9999"), x.get("row_idx") or 0)):
        if w.get("cancelled") or not w.get("txn_date") or not (w.get("amount") or 0):
            continue
        r += 1
        day = as_date(w["txn_date"])
        _put(ws, r, 1, day, num="dd/mm/yyyy", align="center", name="Courier New")
        _put(ws, r, 2, f"{day:%a}", align="center", colour=INK2)
        _put(ws, r, 3, w.get("amount") or 0, num=MONEY, align="right", name="Courier New")
        _put(ws, r, 4, w.get("recipient") or "", align="left")
        _put(ws, r, 5, w.get("entry_type") or "", align="left", colour=INK2, size=9)
        _put(ws, r, 6, w.get("reason") or "", align="left", colour=INK2, size=9)
    r += 1
    _put(ws, r, 1, "Total", bold=True, colour="FFFFFF", fill=BLUE, align="left")
    _put(ws, r, 2, "", fill=BLUE)
    _put(ws, r, 3, f"=SUM(C{first}:C{r - 1})" if r > first else 0, bold=True, num=MONEY,
         align="right", colour="FFFFFF", fill=BLUE, name="Courier New")
    for i in (4, 5, 6):
        _put(ws, r, i, "", fill=BLUE)

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 9
    for i in range(3, 4 + len(days)):
        ws.column_dimensions[get_column_letter(i)].width = 13
    ws.column_dimensions["D"].width = max(ws.column_dimensions["D"].width or 0, 34)
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 26
    _page(ws)
    return ws


# ---------------------------------------------------- range summary (>1 week)
def build_range_summary(wb, sheets, from_date, to_date):
    cols = ["Week ending", "Float"] + OUT_COLS + ["TOTAL OUT", "Cash IN", "Expected close", "Status"]
    n = len(cols)
    ws = wb.create_sheet("Range summary", 0)
    _title(ws, n,
           f"Petty Cash — {len(sheets)} weeks by category · {as_date(from_date):%d %b %Y} – {as_date(to_date):%d %b %Y}",
           f"Generated {datetime.now():%d %b %Y %H:%M}  ·  "
           + ", ".join(s.get("name") for s in sheets))

    hdr = 5
    ws.merge_cells(start_row=hdr, start_column=3, end_row=hdr, end_column=2 + len(CAT_CODES))
    _put(ws, hdr, 3, "VOUCHER CATEGORIES", bold=True, size=9, colour=BLUE, fill=BLUE_PALE, align="center")
    ws.merge_cells(start_row=hdr, start_column=3 + len(CAT_CODES), end_row=hdr, end_column=2 + len(OUT_COLS))
    _put(ws, hdr, 3 + len(CAT_CODES), "SUPPORTING REGISTERS", bold=True, size=9,
         colour=BLUE, fill=BLUE_PALE, align="center")
    ws.merge_cells(start_row=hdr, start_column=3 + len(OUT_COLS), end_row=hdr, end_column=n)
    _put(ws, hdr, 3 + len(OUT_COLS), "TOTALS", bold=True, size=9, colour=BLUE, fill=BLUE_PALE, align="center")
    for c in (1, 2):
        _put(ws, hdr, c, "", fill=BLUE_PALE)
    hdr2 = hdr + 1
    for i, label in enumerate(cols, start=1):
        _put(ws, hdr2, i, label, bold=True, size=9, colour=NAVY, fill=BLUE_LIGHT,
             align="center" if i > 2 else "left", wrap=True)
    ws.row_dimensions[hdr2].height = 30

    r = hdr2
    for sheet in sheets:
        r += 1
        days = week_days(sheet["week_ending"])
        rows = daily_totals(sheet, days)
        agg = {k: sum(rows[day][k] for day in days) for k in OUT_COLS + ["Cash IN"]}
        _put(ws, r, 1, days[-1], num="dd/mm/yyyy", align="center", name="Courier New")
        _put(ws, r, 2, sheet.get("float") or "", align="left")
        for i, key in enumerate(OUT_COLS, start=3):
            _put(ws, r, i, agg[key] or None, num=MONEY, align="right", name="Courier New")
        tot_out = sum(agg[k] for k in OUT_COLS)
        _put(ws, r, 3 + len(OUT_COLS), tot_out or None, bold=True, num=MONEY, align="right", name="Courier New")
        _put(ws, r, 4 + len(OUT_COLS), agg["Cash IN"] or None, num=MONEY, align="right",
             colour=GREEN if agg["Cash IN"] else INK, name="Courier New")
        close = sheet.get("expected_close") or 0
        _put(ws, r, 5 + len(OUT_COLS), close, num=MONEY, align="right", name="Courier New",
             colour=AMBER if close < 0 else INK)
        _put(ws, r, n, sheet.get("status") or "", align="center", size=9,
             colour=GREEN if sheet.get("status") == "Approved" else AMBER)
    first, last = hdr2 + 1, r

    r += 1
    _put(ws, r, 1, "TOTAL", bold=True, colour="FFFFFF", fill=BLUE, align="left")
    _put(ws, r, 2, "", fill=BLUE)
    for i in range(3, 4 + len(OUT_COLS) + 1):
        L = get_column_letter(i)
        _put(ws, r, i, f"=SUM({L}{first}:{L}{last})", bold=True, num=MONEY, align="right",
             colour="FFFFFF", fill=BLUE, name="Courier New")
    for i in (5 + len(OUT_COLS), n):
        _put(ws, r, i, "", fill=BLUE)

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 12
    for i in range(3, n + 1):
        ws.column_dimensions[get_column_letter(i)].width = 13
    ws.freeze_panes = f"C{hdr2 + 1}"
    _page(ws)
    return ws


# ----------------------------------------------------------------- notes tab
def build_notes(wb, sheets, source, checks):
    ws = wb.create_sheet("Notes & checks")
    _put(ws, 1, 1, "VIMIT CONVERTERS LIMITED", bold=True, size=15, colour=NAVY, align="left", box=False)
    _put(ws, 2, 1, "Petty Cash — notes & data-quality checks", bold=True, size=11,
         colour=BLUE, align="left", box=False)
    r = 4
    _put(ws, r, 1, "Basis of preparation", bold=True, colour=NAVY, align="left", box=False)
    for line in [
        f"Source: ERPNext Petty Cash Sheet — {', '.join(s.get('name') for s in sheets)}"
        + (f" ({source})." if source else "."),
        "Weeks run Sunday to Saturday; week_ending is the Saturday.",
        "Cancelled rows are listed on the registers for audit but excluded from every total.",
        "Category codes: " + "; ".join(f"{c} = {l}" for c, l in CATEGORIES) + ".",
        "Cash out = vouchers + parking + bike fuel + forklift + wages/commission + loans issued.",
        "All amounts in KES.",
    ]:
        r += 1
        _put(ws, r, 1, "•  " + line, align="left", box=False, wrap=True)
    r += 2
    _put(ws, r, 1, "Checks", bold=True, colour=NAVY, align="left", box=False)
    if not checks:
        r += 1
        _put(ws, r, 1, "No checks were run.", italic=True, colour=INK2, align="left", box=False)
    for ok, line in checks:
        r += 1
        _put(ws, r, 1, ("✓  " if ok else "!  ") + line, align="left", box=False,
             wrap=True, colour=GREEN if ok else AMBER)
    ws.column_dimensions["A"].width = 150
    ws.sheet_view.showGridLines = False
    return ws


# ---------------------------------------------------------------- entry point
def build_workbook(sheets, *, from_date=None, to_date=None, source="", checks=None):
    """Workbook bytes for one or more Petty Cash Sheet dicts.

    One sheet -> Summary / Vouchers / Registers. More than one -> a Range summary
    tab first (weeks down the rows, categories across the top) then a tab set per
    sheet. Always closes with Notes & checks.
    """
    sheets = sorted(sheets, key=lambda s: (str(s.get("week_ending")), s.get("float") or ""))
    if not sheets:
        raise ValueError("No Petty Cash Sheet in this range.")

    wb = Workbook()
    wb.remove(wb.active)
    multi = len(sheets) > 1
    for sheet in sheets:
        if multi:
            we = as_date(sheet["week_ending"])
            prefix = f"{(sheet.get('float') or 'Float')[:8]} {we:%d-%m} "
        else:
            prefix = ""
        build_summary(wb, sheet, f"{prefix}Summary"[:31])
        build_vouchers(wb, sheet, f"{prefix}Vouchers"[:31])
        build_registers(wb, sheet, f"{prefix}Registers"[:31])
    if multi:
        build_range_summary(wb, sheets,
                            from_date or sheets[0]["week_ending"],
                            to_date or sheets[-1]["week_ending"])
    build_notes(wb, sheets, source, checks or [])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def filename(sheets, from_date=None, to_date=None):
    if len(sheets) == 1:
        we = as_date(sheets[0]["week_ending"])
        fl = (sheets[0].get("float") or "").replace(" ", "-")
        return f"VCL_Petty_Cash_{fl}_WE_{we:%Y-%m-%d}.xlsx" if fl else f"VCL_Petty_Cash_WE_{we:%Y-%m-%d}.xlsx"
    a = as_date(from_date or sheets[0]["week_ending"])
    b = as_date(to_date or sheets[-1]["week_ending"])
    return f"VCL_Petty_Cash_{a:%Y-%m-%d}_to_{b:%Y-%m-%d}.xlsx"
