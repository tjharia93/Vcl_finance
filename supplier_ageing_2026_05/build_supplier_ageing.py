"""May 2026 Supplier Ageing — compare Zoho-received supplier statements vs QBO AP ledger.

Reuses Zoho REST + QBOClient already running in vcl-zoho-mcp / CommandCentre.

Outputs:
  - Supplier_Ageing_May_2026.xlsx   (summary + per-supplier tabs)
  - Statements_Received/*.eml       (raw email with all attachments)
  - Statements_Received/*.pdf/.xlsx (each attachment extracted)
"""
from __future__ import annotations

import email
import json
import re
import sys
import urllib.error
import urllib.request
from datetime import datetime, timezone
from email import policy
from pathlib import Path

# --- import existing helpers ------------------------------------------------
sys.path.insert(0, "/home/tanujharia/projects/vcl-zoho-mcp")
sys.path.insert(0, "/opt/vcl/CommandCentre")

from auth import MAIL_BASE, ZOHO_ACCOUNT_ID, ZOHO_DC, auth_header  # noqa: E402
from connection.qbo_client import qbo_client  # noqa: E402

OUTPUT_DIR = Path("/mnt/vimit/apps/02. FINANCE/2026/05_May/Supplier Ageing")
EML_DIR = OUTPUT_DIR / "Statements_Received"
EML_DIR.mkdir(parents=True, exist_ok=True)

# Statement-as-of date used for QBO snapshot (all April-end statements).
SNAPSHOT_DATE = "2026-04-30"

# 9 supplier statements received May 2026, with their QBO vendor mapping.
# vendor_id=None means no QBO master row exists yet (recon gap to flag).
STATEMENTS = [
    {
        "supplier": "ROK Industries Limited",
        "from": "pa@rokindltd.co.ke",
        "subject": "STATEMENT OF ACCOUNTS AS AT 30.04.2026-VIMIT CONVERTERS LTD",
        "message_id": "1777971709776162600",
        "folder_id": "7661821000000008014",
        "received": "2026-05-05",
        "qbo_vendor_id": "1437",
        "qbo_vendor_name": "ROK INDUSTRIES LIMITED - Supplier",
        "statement_balance": 73988.00,
        "statement_currency": "KES",
        "statement_as_of": "2026-04-30",
        "notes": "Statement source: PDF (Tally-format ledger). Closing balance 73,988.00 KES as at 30-Apr-26.",
    },
    {
        "supplier": "Avery Dennison (Kenya) Pvt. Ltd",
        "from": "ADKENYA.CS@ap.averydennison.com",
        "subject": "Statement for account V00001 as of 04/30/26",
        "message_id": "1778046198889141500",
        "folder_id": "7661821000001292002",
        "received": "2026-05-05",
        "qbo_vendor_id": "1122",
        "qbo_vendor_name": "Avery Dennison (Kenya) Pvt. Ltd",
        "statement_balance": 6720105.66,
        "statement_currency": "KES",
        "statement_as_of": "2026-04-30",
        "notes": "Statement source: PDF (SAP ageing). TOTAL DUE 6,720,105.66 KES (Current 6,551,782.67 + 90-120d 47,717.72 + 120+ 120,605.27).",
    },
    {
        "supplier": "Faha Graphics Enterprises Ltd",
        "from": "admin1@fahagraphics.co.ke",
        "subject": "RE: RE: Statement (as at 30.04.2026)",
        "message_id": "1777977226182141500",
        "folder_id": "7661821000001292002",
        "received": "2026-05-05",
        "qbo_vendor_id": "1094",
        "qbo_vendor_name": "Faha Graphics Enterprises Ltd",
        "statement_balance": 257004.00,
        "statement_currency": "KES",
        "statement_as_of": "2026-04-30",
        "notes": "Statement source: PDF ageing. Amount Due 257,004.00 KES (Current 79,112 + 30d 154,396 + 60d 16,723 + 90d 6,773).",
    },
    {
        "supplier": "Oswald Overseas Corporation",
        "from": "oswaldoverseas@yahoo.com",
        "subject": "Re: Oswald Overseas Corporation Statement",
        "message_id": "1777964374785141500",
        "folder_id": "7661821000001292002",
        "received": "2026-05-05",
        "qbo_vendor_id": "1384",
        "qbo_vendor_name": "OSWALD OVERSEAS CORPORATION",
        "statement_balance": 58000.00,
        "statement_currency": "KES",
        "statement_as_of": "2026-05-05",
        "notes": "Statement source: email body (no attachment). 1 invoice: 10/4/26 INV 21223 KES 58,000.",
    },
    {
        "supplier": "Highway Stores Limited",
        "from": "info@highwaystores.com",
        "subject": "Statement from Highway Stores",
        "message_id": "1777876908612162600",
        "folder_id": "7661821000000008014",
        "received": "2026-05-04",
        "qbo_vendor_id": "1251",
        "qbo_vendor_name": "HIGHWAY STORES LIMITED",
        "statement_balance": 31610.00,
        "statement_currency": "KES",
        "statement_as_of": "2026-04-30",
        "notes": "Statement source: PDF. Amount Due 31,610 KES (Current 0, 30d 8,800, 60-90d 12,370, 90+ 10,440).",
    },
    {
        "supplier": "Tazar Trading",
        "from": "info@tazartrading.co.ke",
        "subject": "Statement from Tazar Trading",
        "message_id": "1777874526204162800",
        "folder_id": "7661821000000008014",
        "received": "2026-05-04",
        "qbo_vendor_id": "1500",
        "qbo_vendor_name": "TAZAR TRADING",
        "statement_balance": 340.00,
        "statement_currency": "KES",
        "statement_as_of": "2026-04-30",
        "notes": "Statement source: PDF. Single line residue INV 16628 KES 340.00 (Orig 37,120).",
    },
    {
        "supplier": "Epitome Lubricants Limited",
        "from": "accounts@epitomelubes.com",
        "subject": "Statement from EPITOME LUBRICANTS LIMITED",
        "message_id": "1777875765911141500",
        "folder_id": "7661821000000008014",
        "received": "2026-05-04",
        "qbo_vendor_id": "1198",
        "qbo_vendor_name": "Epitome Lubricants",
        "statement_balance": 54508.05,
        "statement_currency": "KES",
        "statement_as_of": "2026-04-30",
        "notes": "Statement source: PDF. Amount Due 54,508.05 KES. Bal-fwd 1,744.05 + INV 1909/1948/1983 less VAT WH 1,040.",
    },
    {
        "supplier": "Bahati Industries Limited",
        "from": "accounts@bahati.co.ke",
        "subject": "Statement from Bahati Industries Ltd",
        "message_id": "1777812224681141600",
        "folder_id": "7661821000000008014",
        "received": "2026-05-03",
        "qbo_vendor_id": "1086",
        "qbo_vendor_name": "BAHATI INDUSTRIES LIMITED (+ Construction split 1127)",
        "qbo_extra_vendor_ids": ["1127"],
        "statement_balance": 59453558.60,
        "statement_currency": "KES",
        "statement_as_of": "2026-04-30",
        "notes": "INTERCOMPANY (VCL group). Statement Amount Due 59,453,558.60 KES (90+ past due 55,589,530.16 + 30d 3,864,028.44). Big variance vs QBO — investigate, likely construction invoices also booked to BAHATI INDUSTRIES - CONSTRUCTION (1127).",
    },
    {
        "supplier": "PG Paper Company Ltd",
        "from": "Finance@pgpaper.com",
        "subject": "Statement of Account PG Paper / VIMIT CONVERTERS LIMITED.",
        "message_id": "1778256421546141600",
        "folder_id": "7661821000000008014",
        "received": "2026-05-08",
        "qbo_vendor_id": None,
        "qbo_vendor_name": "(not in QBO master)",
        "statement_balance": 25658.10,
        "statement_currency": "USD",
        "statement_as_of": "2026-05-08",
        "notes": "Statement source: email body. 1 open invoice doc 12093 dated 09/03/2026, due 10/07/2026, Net 120 days, USD 25,658.10. No QBO vendor master row — create supplier in QBO + post the bill.",
    },
]


# --- Zoho .eml + attachment extraction --------------------------------------

def slugify(s: str) -> str:
    return re.sub(r"[^a-zA-Z0-9]+", "_", s.strip().lower()).strip("_")[:50]


def fetch_eml(message_id: str) -> str | None:
    url = f"https://mail.zoho.{ZOHO_DC}/api/accounts/{ZOHO_ACCOUNT_ID}/messages/{message_id}/originalmessage"
    hdr = auth_header()
    try:
        with urllib.request.urlopen(urllib.request.Request(url, headers=hdr), timeout=60) as r:
            payload = json.loads(r.read().decode())
    except urllib.error.HTTPError as e:
        print(f"  [HTTP {e.code}] {message_id}: {e.read().decode()[:200]}", file=sys.stderr)
        return None
    except Exception as e:
        print(f"  [err] {message_id}: {e}", file=sys.stderr)
        return None
    return payload.get("data", {}).get("content")


def save_statement(stmt: dict) -> dict:
    slug = slugify(stmt["supplier"])
    eml_path = EML_DIR / f"{slug}_{stmt['message_id']}.eml"
    attachments: list[str] = []
    if not eml_path.exists() or eml_path.stat().st_size == 0:
        content = fetch_eml(stmt["message_id"])
        if content is None:
            print(f"  ! could not fetch eml for {stmt['supplier']}", file=sys.stderr)
            return {"eml_path": "", "attachments": []}
        eml_path.write_text(content, encoding="utf-8")

    raw = eml_path.read_text(encoding="utf-8", errors="replace")
    msg = email.message_from_string(raw, policy=policy.default)
    for part in msg.walk():
        if part.get_content_disposition() != "attachment":
            continue
        fn = part.get_filename() or "attachment.bin"
        fn_safe = re.sub(r"[^A-Za-z0-9._-]+", "_", fn)
        out_path = EML_DIR / f"{slug}__{fn_safe}"
        try:
            payload = part.get_payload(decode=True)
        except Exception:
            payload = None
        if payload is None:
            continue
        out_path.write_bytes(payload)
        attachments.append(out_path.name)

    return {"eml_path": eml_path.name, "attachments": attachments}


# --- QBO ledger queries -----------------------------------------------------

def _query_for(client, vendor_id: str, snapshot: str) -> dict:
    bills = client.query(
        f"select Id, DocNumber, TxnDate, DueDate, TotalAmt, Balance, CurrencyRef "
        f"from Bill where VendorRef = '{vendor_id}' "
        f"and TxnDate <= '{snapshot}' MAXRESULTS 1000"
    ).get("QueryResponse", {}).get("Bill", [])

    pays = client.query(
        f"select Id, DocNumber, TxnDate, TotalAmt, CurrencyRef "
        f"from BillPayment where VendorRef = '{vendor_id}' "
        f"and TxnDate <= '{snapshot}' MAXRESULTS 1000"
    ).get("QueryResponse", {}).get("BillPayment", [])

    credits = client.query(
        f"select Id, DocNumber, TxnDate, TotalAmt, CurrencyRef "
        f"from VendorCredit where VendorRef = '{vendor_id}' "
        f"and TxnDate <= '{snapshot}' MAXRESULTS 1000"
    ).get("QueryResponse", {}).get("VendorCredit", [])

    vendor = client.query(
        f"select Id, DisplayName, Balance, CurrencyRef from Vendor where Id = '{vendor_id}'"
    ).get("QueryResponse", {}).get("Vendor", [{}])[0]
    return {"bills": bills, "payments": pays, "credits": credits, "vendor": vendor}


def qbo_vendor_summary(client, vendor_id: str, snapshot: str, extra_vendor_ids: list[str] | None = None) -> dict:
    """Return KPIs as of snapshot date.

    Historical net AP at snapshot = bills_total_to_date − payments_total_to_date − credits_total_to_date.
    Optional extra_vendor_ids merges sibling vendor masters (e.g. Bahati Construction split).
    """
    pulls = [_query_for(client, vendor_id, snapshot)]
    for v in (extra_vendor_ids or []):
        pulls.append(_query_for(client, v, snapshot))

    bills = [b for p in pulls for b in p["bills"]]
    pays = [b for p in pulls for b in p["payments"]]
    credits = [b for p in pulls for b in p["credits"]]
    vendor = pulls[0]["vendor"]

    bill_total = sum(float(b.get("TotalAmt") or 0) for b in bills)
    bill_open = sum(float(b.get("Balance") or 0) for b in bills)
    paid_total = sum(float(p.get("TotalAmt") or 0) for p in pays)
    credit_total = sum(float(c.get("TotalAmt") or 0) for c in credits)
    credit_open = credit_total

    # Historical net AP at the snapshot date — the right number to compare against statements.
    historical_open = bill_total - paid_total - credit_total

    # Sum of live balances across all merged vendor masters.
    live_balance = float(vendor.get("Balance") or 0)
    for p in pulls[1:]:
        live_balance += float(p["vendor"].get("Balance") or 0)

    return {
        "bills": bills,
        "payments": pays,
        "credits": credits,
        "bill_total": bill_total,
        "bill_open": bill_open,
        "paid_total": paid_total,
        "credit_total": credit_total,
        "credit_open": credit_open,
        "net_open": historical_open,
        "vendor_balance_live": live_balance,
    }


# --- workbook ---------------------------------------------------------------

def build_workbook(rows: list[dict]) -> Path:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Summary"

    NAVY = PatternFill("solid", fgColor="1F3A8A")
    LIGHT = PatternFill("solid", fgColor="EFF6FF")
    AMBER = PatternFill("solid", fgColor="FEF3C7")
    GREEN = PatternFill("solid", fgColor="DCFCE7")
    RED = PatternFill("solid", fgColor="FEE2E2")
    BOLD_WHITE = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    BOLD = Font(bold=True, name="Calibri", size=11)

    headers = [
        "Supplier",
        "Received (Zoho)",
        "Statement as-of",
        "Statement Balance",
        "Statement Currency",
        "QBO Vendor",
        "QBO Currency",
        "QBO Vendor Balance (live)",
        "Variance (QBO live − Stmt)",
        "QBO Open Bills (sum Balance)",
        "QBO Credits (Total)",
        "QBO Historical Net AP @ 30-Apr-26",
        "Attachments",
        "Notes",
    ]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = NAVY
        cell.font = BOLD_WHITE
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="left")
    ws.row_dimensions[1].height = 32

    for r, row in enumerate(rows, 2):
        ws.cell(row=r, column=1, value=row["supplier"]).font = BOLD
        ws.cell(row=r, column=2, value=row["received"])
        ws.cell(row=r, column=3, value=row.get("statement_as_of") or "")
        ws.cell(row=r, column=4, value=row.get("statement_balance"))
        ws.cell(row=r, column=5, value=row.get("statement_currency") or "")
        ws.cell(row=r, column=6, value=row["qbo_vendor_name"])
        ws.cell(row=r, column=7, value=row.get("qbo_currency") or "")
        # QBO Vendor.Balance returns NEGATIVE for AP we owe — invert so the workbook
        # shows positive balances consistent with the supplier statement convention.
        raw_bal = row.get("vendor_balance_live")
        live_bal = (-raw_bal) if raw_bal is not None else None
        ws.cell(row=r, column=8, value=live_bal)
        # Variance: live AP minus statement (positive = QBO higher than supplier).
        live_var = None
        if live_bal is not None and row.get("statement_balance") is not None:
            live_var = round(live_bal - row["statement_balance"], 2)
        var_cell = ws.cell(row=r, column=9, value=live_var)
        ws.cell(row=r, column=10, value=row.get("bill_open"))
        ws.cell(row=r, column=11, value=row.get("credit_open"))
        ws.cell(row=r, column=12, value=row.get("net_open"))
        ws.cell(row=r, column=13, value="\n".join(row.get("attachments") or []))
        ws.cell(row=r, column=14, value=row.get("notes") or "")

        for col in (4, 8, 9, 10, 11, 12):
            ws.cell(row=r, column=col).number_format = "#,##0.00;[Red]-#,##0.00;-"
            ws.cell(row=r, column=col).alignment = Alignment(horizontal="right")

        if live_var is None or row.get("statement_balance") is None:
            var_cell.fill = AMBER
        elif abs(live_var) < 1:
            var_cell.fill = GREEN
        elif abs(live_var) <= max(500, 0.005 * float(row["statement_balance"] or 0)):
            # Trivial variance (< 500 KES OR < 0.5% of balance) — likely VAT WH timing.
            var_cell.fill = AMBER
        else:
            var_cell.fill = RED

        ws.row_dimensions[r].height = 60
        ws.cell(row=r, column=13).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=r, column=14).alignment = Alignment(wrap_text=True, vertical="top")

    widths = [32, 14, 14, 18, 11, 36, 10, 22, 22, 20, 16, 24, 36, 60]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    # per-supplier tabs
    for row in rows:
        tab = re.sub(r"[^A-Za-z0-9 ]+", "_", row["supplier"])[:28] or "Supplier"
        ws2 = wb.create_sheet(title=tab)

        ws2.cell(row=1, column=1, value=row["supplier"]).font = Font(bold=True, size=14, color="1F3A8A")
        ws2.cell(row=2, column=1, value=f"QBO vendor: {row['qbo_vendor_name']}")
        ws2.cell(row=3, column=1, value=f"Statement received {row['received']} from {row.get('from','')}")
        ws2.cell(row=4, column=1, value=f"QBO snapshot date: {SNAPSHOT_DATE}")
        if row.get("notes"):
            ws2.cell(row=5, column=1, value=f"Notes: {row['notes']}").alignment = Alignment(wrap_text=True)
            ws2.row_dimensions[5].height = 30
        ws2.column_dimensions["A"].width = 18

        if row.get("bills") is None:
            ws2.cell(row=7, column=1, value="No QBO data — vendor not mapped.").font = Font(italic=True, color="B45309")
            continue

        r_idx = 7
        # Bills table
        ws2.cell(row=r_idx, column=1, value="QBO Bills (≤ 30-Apr-26)").font = BOLD
        ws2.cell(row=r_idx, column=1).fill = NAVY
        ws2.cell(row=r_idx, column=1).font = BOLD_WHITE
        r_idx += 1
        bhdr = ["TxnDate", "DueDate", "DocNumber", "TotalAmt", "Open Balance", "Currency"]
        for i, h in enumerate(bhdr, 1):
            cell = ws2.cell(row=r_idx, column=i, value=h)
            cell.fill = LIGHT
            cell.font = BOLD
        r_idx += 1
        for b in sorted(row["bills"], key=lambda x: (x.get("TxnDate") or "")):
            ws2.cell(row=r_idx, column=1, value=b.get("TxnDate"))
            ws2.cell(row=r_idx, column=2, value=b.get("DueDate"))
            ws2.cell(row=r_idx, column=3, value=b.get("DocNumber"))
            ws2.cell(row=r_idx, column=4, value=float(b.get("TotalAmt") or 0)).number_format = "#,##0.00"
            ws2.cell(row=r_idx, column=5, value=float(b.get("Balance") or 0)).number_format = "#,##0.00"
            ws2.cell(row=r_idx, column=6, value=(b.get("CurrencyRef") or {}).get("value"))
            r_idx += 1
        ws2.cell(row=r_idx, column=3, value="Open total").font = BOLD
        ws2.cell(row=r_idx, column=5, value=row["bill_open"]).number_format = "#,##0.00"
        ws2.cell(row=r_idx, column=5).font = BOLD
        r_idx += 2

        # Credits
        if row["credits"]:
            ws2.cell(row=r_idx, column=1, value="QBO Vendor Credits").font = BOLD_WHITE
            ws2.cell(row=r_idx, column=1).fill = NAVY
            r_idx += 1
            chdr = ["TxnDate", "DocNumber", "TotalAmt", "Currency"]
            for i, h in enumerate(chdr, 1):
                cell = ws2.cell(row=r_idx, column=i, value=h)
                cell.fill = LIGHT
                cell.font = BOLD
            r_idx += 1
            for c in sorted(row["credits"], key=lambda x: (x.get("TxnDate") or "")):
                ws2.cell(row=r_idx, column=1, value=c.get("TxnDate"))
                ws2.cell(row=r_idx, column=2, value=c.get("DocNumber"))
                ws2.cell(row=r_idx, column=3, value=float(c.get("TotalAmt") or 0)).number_format = "#,##0.00"
                ws2.cell(row=r_idx, column=4, value=(c.get("CurrencyRef") or {}).get("value"))
                r_idx += 1
            r_idx += 1

        # Payments (recent 24 only to keep readable)
        ws2.cell(row=r_idx, column=1, value="QBO Bill Payments (last 24)").font = BOLD_WHITE
        ws2.cell(row=r_idx, column=1).fill = NAVY
        r_idx += 1
        phdr = ["TxnDate", "DocNumber", "TotalAmt", "Currency"]
        for i, h in enumerate(phdr, 1):
            cell = ws2.cell(row=r_idx, column=i, value=h)
            cell.fill = LIGHT
            cell.font = BOLD
        r_idx += 1
        pays_sorted = sorted(row["payments"], key=lambda x: (x.get("TxnDate") or ""), reverse=True)[:24]
        for p in pays_sorted:
            ws2.cell(row=r_idx, column=1, value=p.get("TxnDate"))
            ws2.cell(row=r_idx, column=2, value=p.get("DocNumber"))
            ws2.cell(row=r_idx, column=3, value=float(p.get("TotalAmt") or 0)).number_format = "#,##0.00"
            ws2.cell(row=r_idx, column=4, value=(p.get("CurrencyRef") or {}).get("value"))
            r_idx += 1

        for col_letter, w in zip("ABCDEF", (12, 12, 22, 18, 18, 10)):
            ws2.column_dimensions[col_letter].width = w
        ws2.freeze_panes = "A8"

    out = OUTPUT_DIR / f"Supplier_Ageing_May_2026.xlsx"
    wb.save(out)
    return out


def main() -> int:
    client = qbo_client()
    rows: list[dict] = []
    for stmt in STATEMENTS:
        print(f"==> {stmt['supplier']}")
        # 1. download .eml + extract attachments
        saved = save_statement(stmt)
        # 2. QBO ledger
        if stmt.get("qbo_vendor_id"):
            try:
                qbo = qbo_vendor_summary(
                    client,
                    stmt["qbo_vendor_id"],
                    SNAPSHOT_DATE,
                    extra_vendor_ids=stmt.get("qbo_extra_vendor_ids") or [],
                )
            except Exception as e:
                print(f"   ! QBO query failed: {e}", file=sys.stderr)
                qbo = {"bills": [], "payments": [], "credits": [], "bill_total": 0, "bill_open": 0,
                       "paid_total": 0, "credit_total": 0, "credit_open": 0, "net_open": 0}
        else:
            qbo = {"bills": None, "payments": None, "credits": None, "bill_total": None,
                   "bill_open": None, "paid_total": None, "credit_total": None,
                   "credit_open": None, "net_open": None}

        cur = None
        if qbo.get("bills"):
            cur = (qbo["bills"][0].get("CurrencyRef") or {}).get("value")
        elif qbo.get("payments"):
            cur = (qbo["payments"][0].get("CurrencyRef") or {}).get("value")

        net_open = qbo.get("net_open")
        stmt_bal = stmt.get("statement_balance")
        variance = None
        # Compare against QBO net AP from historical bill balances at snapshot date.
        if net_open is not None and stmt_bal is not None:
            variance = round(net_open - stmt_bal, 2)

        rows.append({
            **stmt,
            **qbo,
            "qbo_currency": cur,
            "variance": variance,
            "attachments": saved["attachments"],
        })

    out = build_workbook(rows)
    print(f"\nWorkbook: {out}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
